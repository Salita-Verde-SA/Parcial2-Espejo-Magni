from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from fastapi import HTTPException
from fastapi.responses import StreamingResponse

from app.core.uow import UnitOfWork
from app.modules.productos.model import (
    Producto,
    ProductoCreate,
    ProductoPublic,
    ProductoUpdate,
    StockUpdate,
    DisponibilidadUpdate,
    PaginatedProductos,
    IngredienteResumen,
    UnidadMedidaResumen,
)


def calcular_stock_producto(producto_id: int, uow: UnitOfWork) -> int:
    """
    Calcula el stock de un producto usando la fórmula:
    stock_producto = min(floor(stock_ingrediente / cantidad_necesaria))
    
    IMPORTANTE: Esta función debe llamarse DENTRO de un contexto 'with uow:'
    para evitar errores de sesiones de SQLAlchemy.
    
    - Si no tiene ingredientes, retorna 0
    - Si un ingrediente tiene stock 0 o null, el resultado será 0
    - Si cantidad es null o 0, usa default = 1
    """
    # Ya no hace 'with uow' - debe llamarse dentro de un contexto existente
    stocks_data = uow.productos.get_ingrediente_stocks(producto_id)
    if not stocks_data:
        return 0
    
    # Calcular: min(floor(stock_ingrediente / cantidad_necesaria))
    max_products = []
    for _, stock, cantidad_necesaria in stocks_data:
        if stock is None or stock <= 0 or cantidad_necesaria is None or cantidad_necesaria <= 0:
            # Si algún ingrediente no tiene stock o cantidad inválida, no se puede producir nada
            max_products.append(0)
        else:
            productos_posibles = int(stock // cantidad_necesaria)
            max_products.append(productos_posibles)
    
    return min(max_products)


def validar_stock_ingredientes(ingredientes: list, uow: UnitOfWork) -> None:
    """
    Valida que cada ingrediente tenga stock suficiente para la cantidad necesaria.
    Lanza HTTPException si algún ingrediente no tiene stock suficiente.
    
    IMPORTANTE: Esta función debe llamarse DENTRO de un contexto 'with uow:'
    para evitar errores de sesiones de SQLAlchemy.
    
    ingredientes: lista de dicts con keys: ingrediente_id, cantidad
    """
    if not ingredientes:
        return
    
    # Obtener IDs de ingredientes a validar
    ing_ids = [ing.ingrediente_id for ing in ingredientes]
    
    # Ya no hace 'with uow' - debe llamarse dentro de un contexto existente
    # Obtener stock de los ingredientes
    from app.modules.ingredientes.model import Ingrediente
    from sqlmodel import select
    stmt = select(Ingrediente.id, Ingrediente.nombre, Ingrediente.stock_cantidad).where(
        Ingrediente.id.in_(ing_ids),
        Ingrediente.deleted_at.is_(None)
    )
    ingredientes_db = {i.id: (i.nombre, i.stock_cantidad or 0) for i in uow.session.exec(stmt).all()}

    # Validar cada ingrediente
    errores = []
    for ing in ingredientes:
        ing_id = ing.ingrediente_id
        cantidad_necesaria = float(ing.cantidad) if ing.cantidad and float(ing.cantidad) > 0 else 1.0
        
        if ing_id not in ingredientes_db:
            errores.append(f"Ingrediente con ID {ing_id} no encontrado o inactivo")
            continue
            
        nombre, stock_actual = ingredientes_db[ing_id]
        
        if stock_actual < cantidad_necesaria:
            errores.append(
                f"Stock insuficiente para '{nombre}': tienes {stock_actual} pero necesitas {cantidad_necesaria}"
            )
    
    if errores:
        raise HTTPException(
            status_code=400,
            detail="Error de validación: " + "; ".join(errores)
        )


def _enrich(producto: Producto, uow: UnitOfWork) -> ProductoPublic:
    cats = uow.productos.get_categoria_ids(producto.id)
    ings_raw = uow.productos.get_ingredientes(producto.id)
    ings = [IngredienteResumen(**i) for i in ings_raw]
    # Stock calculado como mínimo de los ingredientes
    stock_calculado = calcular_stock_producto(producto.id, uow)
    
    # Obtener unidad de venta si existe
    unidad_venta = None
    if producto.unidad_venta_id:
        unid = uow.unidades.get_by_id(producto.unidad_venta_id)
        if unid:
            unidad_venta = UnidadMedidaResumen(
                id=unid.id,
                nombre=unid.nombre,
                simbolo=unid.simbolo,
            )
    
    # Si el stock calculado es 0, forzar disponible=False
    # Si stock > 0, mantener el valor original de disponible
    disponible_final = producto.disponible if stock_calculado > 0 else False
    
    return ProductoPublic(
        id=producto.id,
        nombre=producto.nombre,
        descripcion=producto.descripcion,
        precio_base=producto.precio_base,
        unidad_venta_id=producto.unidad_venta_id,
        unidad_venta=unidad_venta,
        stock_cantidad=stock_calculado,
        disponible=disponible_final,
        imagen_url=producto.imagen_url,
        created_at=producto.created_at,
        deleted_at=producto.deleted_at,
        categorias=cats,
        ingredientes=ings,
    )


def list_productos(
    nombre: str,
    categoria_id: int | None,
    disponible: str,
    page: int,
    page_size: int,
    uow: UnitOfWork,
) -> PaginatedProductos:
    with uow:
        items, total = uow.productos.list_filtered(nombre, categoria_id, disponible, page, page_size)
        enriched = [_enrich(p, uow) for p in items]
        pages = max(1, -(-total // page_size))
        return PaginatedProductos(items=enriched, total=total, page=page, page_size=page_size, pages=pages)


def list_productos_all(uow: UnitOfWork) -> PaginatedProductos:
    """List all productos including deleted ones (for admin management)."""
    with uow:
        items = uow.productos.get_all()
        enriched = [_enrich(p, uow) for p in items]
        return PaginatedProductos(items=enriched, total=len(items), page=1, page_size=len(items), pages=1)


def get_producto(producto_id: int, uow: UnitOfWork) -> ProductoPublic:
    with uow:
        p = uow.productos.get_by_id_active(producto_id)
        if not p:
            raise HTTPException(status_code=404, detail="Producto no encontrado")
        return _enrich(p, uow)


def create_producto(data: ProductoCreate, uow: UnitOfWork) -> ProductoPublic:
    with uow:
        # Validar stock de ingredientes antes de crear el producto
        if data.ingredientes:
            validar_stock_ingredientes(data.ingredientes, uow)
        
        p = Producto(
            nombre=data.nombre,
            descripcion=data.descripcion,
            precio_base=data.precio_base,
            unidad_venta_id=data.unidad_venta_id,
            disponible=data.disponible,
            imagen_url=data.imagen_url,
        )
        saved = uow.productos.add(p)
        if data.categoria_ids:
            uow.productos.set_categorias(saved.id, data.categoria_ids)
        if data.ingredientes:
            # Convertir lista de IngredienteCantidadInput a lista de dicts
            ingredientes_list = [
                {
                    "ingrediente_id": ing.ingrediente_id,
                    "cantidad": ing.cantidad,
                    "unidad_medida_id": ing.unidad_medida_id,
                    "es_removible": ing.es_removible,
                }
                for ing in data.ingredientes
            ]
            uow.productos.set_ingredientes(saved.id, ingredientes_list)
        return _enrich(saved, uow)


def update_producto(producto_id: int, data: ProductoUpdate, uow: UnitOfWork) -> ProductoPublic:
    with uow:
        p = uow.productos.get_by_id_active(producto_id)
        if not p:
            raise HTTPException(status_code=404, detail="Producto no encontrado")
        
        # Validar stock de ingredientes si se están modificando
        if data.ingredientes is not None:
            validar_stock_ingredientes(data.ingredientes, uow)
        
        if data.nombre is not None:
            p.nombre = data.nombre
        if data.descripcion is not None:
            p.descripcion = data.descripcion
        if data.precio_base is not None:
            p.precio_base = data.precio_base
        if data.unidad_venta_id is not None:
            p.unidad_venta_id = data.unidad_venta_id
        if data.disponible is not None:
            p.disponible = data.disponible
        if data.imagen_url is not None:
            p.imagen_url = data.imagen_url
        p.updated_at = datetime.now(timezone.utc)
        saved = uow.productos.update(p)
        if data.categoria_ids is not None:
            uow.productos.set_categorias(saved.id, data.categoria_ids)
        if data.ingredientes is not None:
            # Convertir lista de IngredienteCantidadInput a lista de dicts
            ingredientes_list = [
                {
                    "ingrediente_id": ing.ingrediente_id,
                    "cantidad": ing.cantidad,
                    "unidad_medida_id": ing.unidad_medida_id,
                    "es_removible": ing.es_removible,
                }
                for ing in data.ingredientes
            ]
            uow.productos.set_ingredientes(saved.id, ingredientes_list)
        return _enrich(saved, uow)


def update_stock(producto_id: int, data: StockUpdate, uow: UnitOfWork) -> ProductoPublic:
    with uow:
        p = uow.productos.get_by_id_active(producto_id)
        if not p:
            raise HTTPException(status_code=404, detail="Producto no encontrado")
        p.stock_cantidad = data.stock_cantidad
        p.disponible = data.disponible
        p.updated_at = datetime.now(timezone.utc)
        saved = uow.productos.update(p)
        return _enrich(saved, uow)


def update_disponibilidad(producto_id: int, data: DisponibilidadUpdate, uow: UnitOfWork) -> ProductoPublic:
    with uow:
        p = uow.productos.get_by_id_active(producto_id)
        if not p:
            raise HTTPException(status_code=404, detail="Producto no encontrado")
        p.disponible = data.disponible
        p.updated_at = datetime.now(timezone.utc)
        saved = uow.productos.update(p)
        return _enrich(saved, uow)


def delete_producto(producto_id: int, uow: UnitOfWork) -> None:
    with uow:
        p = uow.productos.get_by_id_active(producto_id)
        if not p:
            # Verificar si existe pero ya está eliminado
            existente = uow.productos.get_by_id(producto_id)
            if existente and existente.deleted_at is not None:
                raise HTTPException(status_code=400, detail="El producto ya está inactivo")
            raise HTTPException(status_code=404, detail="Producto no encontrado")
        uow.productos.soft_delete(p)


def activate_producto(producto_id: int, uow: UnitOfWork) -> ProductoPublic:
    """Reactiva un producto previamente desactivado."""
    with uow:
        p = uow.productos.get_by_id(producto_id)
        if not p:
            raise HTTPException(status_code=404, detail="Producto no encontrado")
        if p.deleted_at is None:
            raise HTTPException(status_code=400, detail="El producto ya está activo")
        
        # Verificar que las categorías asociadas estén activas
        categoria_ids = uow.productos.get_categoria_ids(producto_id)
        for cat_id in categoria_ids:
            cat = uow.categorias.get_by_id(cat_id)
            if cat and cat.deleted_at is not None:
                cat_nombre = cat.nombre or f"ID {cat_id}"
                raise HTTPException(
                    status_code=400,
                    detail=f"No se puede activar: la categoría '{cat_nombre}' está inactiva",
                )
        
        uow.productos.activate(p)
        return _enrich(p, uow)


def export_excel(uow: UnitOfWork) -> StreamingResponse:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    with uow:
        productos = uow.productos.get_all_active()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ["ID", "Nombre", "Descripción", "Precio Base", "Stock", "Disponible", "Fecha Alta"]
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, p in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=p.id)
        ws.cell(row=row, column=2, value=p.nombre)
        ws.cell(row=row, column=3, value=p.descripcion or "")
        ws.cell(row=row, column=4, value=float(p.precio_base))
        ws.cell(row=row, column=5, value=p.stock_cantidad)
        ws.cell(row=row, column=6, value="Sí" if p.disponible else "No")
        ws.cell(row=row, column=7, value=p.created_at.strftime("%d/%m/%Y"))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=productos.xlsx"},
    )