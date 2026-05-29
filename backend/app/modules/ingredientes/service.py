import io
import math
from datetime import datetime, timezone
from typing import Optional

from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse

from app.core.uow import UnitOfWork
from app.modules.ingredientes.model import (
    Ingrediente,
    IngredienteCreate,
    IngredientePublic,
    IngredienteUpdate,
    PaginatedIngredientes,
)


class IngredienteService:

    def __init__(self, uow: UnitOfWork):
        self.uow = uow

    def list_filtered(
        self,
        nombre:      Optional[str],
        es_alergeno: Optional[bool],
        page:        int,
        page_size:   int,
    ) -> PaginatedIngredientes:
        items, total = self.uow.ingredientes.list_filtered(
            nombre, es_alergeno, page, page_size
        )
        pages = max(1, math.ceil(total / page_size))
        return PaginatedIngredientes(
            items=items,
            total=total,
            page=page,
            page_size=page_size,
            pages=pages,
        )

    def list_all(self) -> PaginatedIngredientes:
        items = self.uow.ingredientes.get_all()
        public_items = [
            IngredientePublic(
                id=ing.id,
                nombre=ing.nombre,
                descripcion=ing.descripcion,
                es_alergeno=ing.es_alergeno,
                es_terminado=ing.es_terminado,
                stock_cantidad=ing.stock_cantidad,
                costo_unitario=ing.costo_unitario,
                created_at=ing.created_at,
                updated_at=ing.updated_at,
                deleted_at=ing.deleted_at,
            )
            for ing in items
        ]
        return PaginatedIngredientes(
            items=public_items,
            total=len(items),
            page=1,
            page_size=len(items),
            pages=1,
        )

    def get_by_id(self, ingrediente_id: int) -> Ingrediente:
        ingrediente = self.uow.ingredientes.get_active_by_id(ingrediente_id)
        if not ingrediente:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Ingrediente no encontrado",
            )
        return ingrediente

    def create(self, ing_in: IngredienteCreate) -> Ingrediente:
        if self.uow.ingredientes.get_by_nombre(ing_in.nombre):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Ya existe un ingrediente con ese nombre",
            )
        ingrediente = Ingrediente(
            nombre=ing_in.nombre,
            descripcion=ing_in.descripcion,
            es_alergeno=ing_in.es_alergeno,
            stock_cantidad=ing_in.stock_cantidad,
        )
        return self.uow.ingredientes.add(ingrediente)

    def update(self, ingrediente_id: int, ing_in: IngredienteUpdate) -> Ingrediente:
        ingrediente = self.uow.ingredientes.get_active_by_id(ingrediente_id)
        if not ingrediente:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Ingrediente no encontrado",
            )

        update_data = ing_in.model_dump(exclude_unset=True)

        if "nombre" in update_data and self.uow.ingredientes.exists_nombre_excluding(
            update_data["nombre"], ingrediente_id
        ):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Ya existe un ingrediente con ese nombre",
            )

        for key, value in update_data.items():
            setattr(ingrediente, key, value)
        ingrediente.updated_at = datetime.now(timezone.utc)

        return self.uow.ingredientes.update(ingrediente)

    def soft_delete(self, ingrediente_id: int) -> None:
        ingrediente = self.uow.ingredientes.get_active_by_id(ingrediente_id)
        if not ingrediente:
            existente = self.uow.ingredientes.get_by_id(ingrediente_id)
            if existente and existente.deleted_at is not None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="El ingrediente ya está inactivo",
                )
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Ingrediente no encontrado",
            )
        self.uow.ingredientes.soft_delete(ingrediente)

    def activate(self, ingrediente_id: int) -> Ingrediente:
        ingrediente = self.uow.ingredientes.get_by_id(ingrediente_id)
        if not ingrediente:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Ingrediente no encontrado",
            )
        if ingrediente.deleted_at is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El ingrediente ya está activo",
            )
        return self.uow.ingredientes.activate(ingrediente)

    def export_excel(self) -> StreamingResponse:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="openpyxl no está instalado",
            )

        ingredientes = self.uow.ingredientes.get_all_active()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ingredientes"

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        headers = ["ID", "Nombre", "Descripción", "Es Alérgeno", "Fecha de Alta", "Últ. Actualización"]
        col_widths = [6, 30, 40, 14, 22, 22]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = width

        ws.row_dimensions[1].height = 20

        for row_idx, ing in enumerate(ingredientes, 2):
            ws.cell(row=row_idx, column=1, value=ing.id).alignment = center_align
            ws.cell(row=row_idx, column=2, value=ing.nombre)
            ws.cell(row=row_idx, column=3, value=ing.descripcion or "")
            ws.cell(row=row_idx, column=4, value="Sí" if ing.es_alergeno else "No").alignment = center_align
            ws.cell(row=row_idx, column=5, value=ing.created_at.strftime("%d/%m/%Y %H:%M"))
            ws.cell(row=row_idx, column=6, value=ing.updated_at.strftime("%d/%m/%Y %H:%M"))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=ingredientes.xlsx"},
        )
